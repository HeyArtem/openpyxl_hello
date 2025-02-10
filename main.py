import os
import sys
from openpyxl import load_workbook

FOLDER = "data"  # Директория для рабочих файлов
FOLDER_RESULT = "result"  # Директория для обработанных файлов

FILENAME = "export.xlsx"  # Имя начального xlsx-файла
ADDED_FILENAME = "добавить.txt"  # Имя файла из которого добавлять данные

FILE_PATH = os.path.join(FOLDER, FILENAME)  # Путь до дир. с xlsx-файлом

ROUNDING_VALUE = 2  # величина на которую нужно увеличить значение
TARGET_CODE = 421730  # Код удаляемой строки


def creating_working_directories():
    """
    Проверка и создадние по необходимости рабочих
    директорий result & data
    """
    # Проверяем, есть ли директории
    if not os.path.exists(FOLDER):
        print(f"[!] Директория: '{FOLDER}' не найдена. Создаю...")
        os.makedirs(FOLDER)

    if not os.path.exists(FOLDER_RESULT):
        print(f"[!] Директория: '{FOLDER_RESULT}' не найдена. Создаю...")
        os.makedirs(FOLDER_RESULT)

    if not os.path.exists(FILE_PATH):
        print(
            f"[!] Пожалуйста положите файлы: '{FILENAME}' & {ADDED_FILENAME} в директорию: '{FOLDER}'\n"
            "Запустите скрипт снова, когда положите туда фаилы.")
        sys.exit()

    return increase()


def increase():
    """
    Если ячейка не пустая, !=0,
    функция округляет число до целого и
    у всех четных чисел увеличивает значение на ROUNDING_VALUE
    """

    # Загружаем Excel
    wb = load_workbook(FILE_PATH)  # Открываю файл
    sheet = wb.active  # Беру активный лист

    for row in sheet.iter_rows(min_row=1, values_only=False):
        # Получаю значение из столбца F (это будет 5-й индекс)
        cell = row[5]  # Беру саму ячейку
        osttehno_value = cell.value

        if osttehno_value is not None:

            if isinstance(osttehno_value, (int, float)):
                new_value = round(osttehno_value)
                if new_value % 2 == 0 and new_value != 0:
                    new_value += ROUNDING_VALUE
                    cell.value = new_value
            else:
                print(f"[!] В ячейке F{row[0].row} содержится некорректное значение: {osttehno_value}. Пропускаем.")

    # Сохраня изменения в новый файл
    wb.save(f"{FOLDER_RESULT}/increase.xlsx")
    wb.close()  # Закрываем книгу
    print("[✔] У всех четных позиций значение увеличено на 2\n"
          f"Результат сохранен в {FOLDER_RESULT}/'increase.xlsx'\n")
    return adding_data()


def adding_data():
    """
    Добавляет данные из txt-файла в xlsx-файл,
    формат ячейки меняет на число
    """

    file_path = os.path.join(FOLDER, ADDED_FILENAME)

    # Проверка наличия файла с добавляемыми данными (ADDED_FILENAME)
    if not os.path.exists(file_path):
        print(
            f"[!] Файл '{ADDED_FILENAME}' не найден в '{FOLDER}'!\n"
            "Запустите скрипт снова, когда положите туда фаилы."
        )
        return

    parsed_data = []

    with open(file_path, "r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()

            if set(line) == {"-"} or line == "":
                continue

            parts = [p.strip() for p in line.split("|") if p.strip()]

            if len(parts) < 10:  # Проверяю, чтобы строка имела нужное количество данных
                print(f"[!] Ошибка парсинга строки: {line}")
                continue

            parsed_data.append(parts)

    if not parsed_data:
        print("[!] Нет данных для добавления!")
        return

    wb = load_workbook(f"{FOLDER_RESULT}/increase.xlsx")
    sheet = wb.active

    # Определяю последнюю заполненную строку (без пустых)
    last_row = sheet.max_row
    while last_row > 0 and all(cell.value is None for cell in sheet[last_row]):
        last_row -= 1  # Ищю последнюю строку с данными

    # добавляю данные в виде числа
    # Колонки, которые должны быть числами (A, C, D, E, F, G, H, I)
    numeric_columns = {1, 3, 4, 5, 6, 7, 8, 9}  # Индексы колонок

    for row_data in parsed_data:
        last_row += 1

        for col, value in enumerate(row_data, start=1):
            if col in numeric_columns:
                try:
                    value = float(value.replace(",", "."))  # Заменяю запятую на точку и конвертирую
                except ValueError:
                    pass  # Если не получилось преобразовать, оставляю как есть

            sheet.cell(row=last_row, column=col, value=value)

    # Сохраняю изменения
    result_file = os.path.join(FOLDER_RESULT, "updated.xlsx")
    wb.save(result_file)
    wb.close()

    print(f"[✔] Новые данные добавлены. Сохранено в '{result_file}'\n")
    return delete_row()


def delete_row():
    """
    Функция удаляет строку из xlsx-файла
    по коду (TARGET_CODE) из столбика А
    """

    wb = load_workbook(f"{FOLDER_RESULT}/updated.xlsx")

    sheet = wb.active

    target_row = None  # Сюда запишем номер строки

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=False):
        if row[0].value == TARGET_CODE:
            target_row = row[0].row
            break

    if target_row:
        sheet.delete_rows(target_row)
        # Сохраняем результат
        result_file = os.path.join(FOLDER_RESULT, "without_row.xlsx")
        wb.save(result_file)
        wb.close()
        print(f"[✔] Строка с кодом {TARGET_CODE} (№{target_row}) удалена.")
        print(f"[✔] Файл с удаленной строкой сохранён в '{result_file}'\n")
        return concatenation()
    else:
        print(f"[!] Код {TARGET_CODE} не найден!")
        wb.close()
        return concatenation()


def concatenation():
    """
    Функция находит строки содержащие "10"
    конкатенирует к ним цену
    """

    file_path = os.path.join(FOLDER_RESULT, "without_row.xlsx")
    wb = load_workbook(file_path)
    sheet = wb.active

    # Проходим по строкам
    for row in sheet.iter_rows(min_row=2, values_only=False):  # min_row=2, чтобы пропустить заголовки
        name_cell = row[1]  # Столбец B (name)
        price_cell = row[2]  # Столбец C (cena)

        if name_cell.value and "10" in str(name_cell.value):  # Ищу наличие "10" в имени
            name_cell.value = f"{name_cell.value} ({price_cell.value})"

    result_file = os.path.join(FOLDER_RESULT, "concatenated.xlsx")
    wb.save(result_file)
    wb.close()

    print(f"[✔] Конкатенация завершена. Результат сохранён в '{result_file}'\n")
    return sorting_data()


def sorting_data():
    """
    Функция сортирует данные по столбцу С
    """

    file_path = os.path.join(FOLDER_RESULT, "concatenated.xlsx")
    wb = load_workbook(file_path)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    # Сортирую по третьему столбцу (cena)
    data.sort(key=lambda x: x[2] if isinstance(x[2], (int, float)) else float('inf'))

    # Перезаписываю отсортированные данные
    for idx, row in enumerate(data, start=2):  # Начинаем со 2-й строки, т.к. 1-я — заголовки
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)

    result_file = os.path.join(FOLDER_RESULT, "sorting_data.xlsx")
    wb.save(result_file)
    wb.close()

    print(f"[✔] Сортировка завершена. Результат сохранён в '{result_file}'\n")


def main():
    creating_working_directories()


if __name__ == '__main__':
    main()
